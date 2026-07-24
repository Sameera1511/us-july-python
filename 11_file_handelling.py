# File Handelling :

    # Using python we can create any type of 
    # files.

    # Even we can Read, Write, Update files
    # using python.

    # We also can delete files using some small
    # yet very important built in modules.

#--------------------------------------------------------------

# Create a file.

# file1 = open("./my_files/sample.txt", "w")
# file2 = open("./my_files/file1.txt", "w")
# file3 = open("./my_files/code.py", "w")


# Delete a file:

import os

# os.remove("./my_files/sample.txt")

# if os.path.exists("./my_files/code.py"):
#     os.remove("./my_files/code.py")
# else:
#     print("./my_files/code.py does not exists.")

# os.remove("./my_files/file1.txt")

#-------------------------------------

# Create and write into a file:
# Mode: "w" => creates file and overwrites a content into it.

# file5 = open("./my_files/file1.txt", "w")
# file5.write("This is python's file handeling..!")
# file5.close()

# file6 = open("./my_files/file1.txt", "w")
# file6.write("Evanka Trump is one of the most intelligent women on earth..")
# file6.close()

#----------------------------------

# file7 = open("./my_files/file1.txt", "a")
# file7.write("\nPython is awesome..!")
# file7.close()

#-------------------------------------

# Read a file:

# file1 = open("./my_files/file1.txt", "r")
# content = file1.read()
# print(content)

#---------------------------------------------

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "sheet1"

ws["A1"] = "Firstname"
ws["B1"] = "Lastname"
ws["C1"] = "Age"
ws["D1"] = "Salary"
ws["E1"] = "Email"

data = [
    ["Ajain", "Laxmi", "Dev", "Jay", "Imran", "Rameeza"],
    ["Kolas", "Kurmi", "Patel", "Shah", "Khan", "Sait"],
    [23, 12, 10, 20, 22, 18],
    [4000, 4200, 2100, 5000, 4500, 3600],
    ["ajain123@gmail.com", "laxmi456@outlook.com", "dev111@microsoft.com", "jay@google.com", "imran1@ibm.com", "rameeza.sait@nvidia.com"]
]

for row in data:
    ws.append(row)

ws["C10"] = "Dev Patel"
ws["C10"] = "Ajain Kolas"

wb.save("./my_files/output.xlsx")
